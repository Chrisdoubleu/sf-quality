from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl

CANON = {
    "1": "Quality Authority - Enterprise (Full)",
    "9": "Quality Workflow Approver - Plant",
    "10": "Quality Workflow Coordinator",
    "13": "Floor Supervisor - Initiate & Manage",
    "14": "Floor Reporter - Initiate Only",
}


def ensure_col(ws, name: str) -> int:
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if name in headers:
        return headers.index(name) + 1
    idx = ws.max_column + 1
    ws.cell(1, idx, name)
    return idx


def patch_dim_role(wb: openpyxl.Workbook):
    ws = wb["DIM_Role"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    i_id = headers.index("RoleId") + 1
    i_name = headers.index("RoleName") + 1
    i_canon = ensure_col(ws, "RoleName_Canonical")
    i_legacy = ensure_col(ws, "RoleName_Source")

    for r in range(2, ws.max_row + 1):
        rid = str(ws.cell(r, i_id).value)
        current = ws.cell(r, i_name).value
        if rid in CANON:
            ws.cell(r, i_legacy, current)
            ws.cell(r, i_name, CANON[rid])
            ws.cell(r, i_canon, CANON[rid])


def patch_role_refs(wb: openpyxl.Workbook, sheet_name: str):
    ws = wb[sheet_name]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if "RequiredRoleId" not in headers or "RequiredRole_Ref" not in headers:
        return

    i_id = headers.index("RequiredRoleId") + 1
    i_ref = headers.index("RequiredRole_Ref") + 1
    i_canon = ensure_col(ws, "RequiredRoleName_Canonical")
    i_source = ensure_col(ws, "RequiredRoleLabel_Source")

    for r in range(2, ws.max_row + 1):
        rid = ws.cell(r, i_id).value
        if rid in (None, ""):
            continue
        rid_s = str(rid)
        if rid_s not in CANON:
            continue
        current = ws.cell(r, i_ref).value
        ws.cell(r, i_source, current)
        ws.cell(r, i_ref, CANON[rid_s])
        ws.cell(r, i_canon, CANON[rid_s])


def write_xref_sheet(wb: openpyxl.Workbook):
    name = "ROLE_CANONICAL_XREF"
    if name in wb.sheetnames:
        ws = wb[name]
        wb.remove(ws)
    ws = wb.create_sheet(name)
    ws.append(["RoleId", "CanonicalRoleName", "LegacyOrAliasLabel", "SourceSheet"])

    dim = wb["DIM_Role"]
    headers = [dim.cell(1, c).value for c in range(1, dim.max_column + 1)]
    i_id = headers.index("RoleId") + 1
    i_src = headers.index("RoleName_Source") + 1 if "RoleName_Source" in headers else None

    for r in range(2, dim.max_row + 1):
        rid = str(dim.cell(r, i_id).value)
        if rid not in CANON:
            continue
        legacy = dim.cell(r, i_src).value if i_src else None
        ws.append([rid, CANON[rid], legacy, "DIM_Role"])


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--in", dest="input_path", required=True)
    parser.add_argument("--out", dest="output_path", required=True)
    args = parser.parse_args()

    inp = Path(args.input_path)
    out = Path(args.output_path)

    wb = openpyxl.load_workbook(inp)
    patch_dim_role(wb)
    for sheet_name in ("FACT_Transition", "PATH_STEP"):
        if sheet_name in wb.sheetnames:
            patch_role_refs(wb, sheet_name)
    write_xref_sheet(wb)
    wb.save(out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
