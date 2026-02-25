import unittest
from pathlib import Path

import openpyxl


class PathmapRoleTests(unittest.TestCase):
    def test_dim_role_uses_canonical_names(self):
        p = Path("docs/Quality-Event Classification/Quality_Triage_PathMap.xlsx")
        self.assertTrue(p.exists(), "workbook missing")
        wb = openpyxl.load_workbook(p, data_only=True)
        ws = wb["DIM_Role"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        i_id = headers.index("RoleId") + 1
        i_name = headers.index("RoleName") + 1

        expected = {
            "1": "Quality Authority - Enterprise (Full)",
            "9": "Quality Workflow Approver - Plant",
            "10": "Quality Workflow Coordinator",
            "13": "Floor Supervisor - Initiate & Manage",
            "14": "Floor Reporter - Initiate Only",
        }

        bad = []
        for r in range(2, ws.max_row + 1):
            rid = str(ws.cell(r, i_id).value)
            if rid not in expected:
                continue
            name = str(ws.cell(r, i_name).value)
            if name != expected[rid]:
                bad.append((rid, name, expected[rid]))
        self.assertEqual([], bad)


if __name__ == "__main__":
    unittest.main()
